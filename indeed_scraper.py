"""
Indeed Python Jobs Scraper
==========================
Playwright-based scraper that extracts Python Developer jobs from Indeed
and exports them to CSV / Excel.

Note: Indeed frequently changes selectors and uses strong anti-bot measures.
This script is designed for educational / personal use. Use responsibly and
respect robots.txt + terms of service.

Author: Ahmad Raza
"""

import asyncio
import logging
from pathlib import Path

import pandas as pd
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-7s | %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("indeed-scraper")

# ---------------- CONFIG ----------------
SEARCH_URL = "https://www.indeed.com/jobs?q=python+developer&l="
HEADLESS = True
MAX_PAGES = 3          # safety limit
OUTPUT_DIR = Path("output")
OUTPUT_CSV = OUTPUT_DIR / "indeed_python_jobs.csv"
OUTPUT_XLSX = OUTPUT_DIR / "indeed_python_jobs.xlsx"


async def scrape_indeed():
    OUTPUT_DIR.mkdir(exist_ok=True)
    jobs = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=HEADLESS,
            args=["--disable-blink-features=AutomationControlled", "--no-sandbox"],
        )
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/140.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1366, "height": 768},
        )
        # Stealth
        await context.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        )
        page = await context.new_page()

        for page_num in range(MAX_PAGES):
            start = page_num * 10
            url = f"{SEARCH_URL}&start={start}"
            log.info("Page %d → %s", page_num + 1, url)

            try:
                await page.goto(url, wait_until="domcontentloaded", timeout=45000)
                await page.wait_for_timeout(2500)

                # Job cards (Indeed changes class names often)
                cards = await page.locator("div.job_seen_beacon, div.cardOutline, li.css-5lfssg").all()
                if not cards:
                    cards = await page.locator("[data-jk]").all()

                log.info("Cards found: %d", len(cards))
                if not cards:
                    log.warning("No cards on this page – stopping")
                    break

                for card in cards:
                    try:
                        title_el = card.locator("h2 a, h2 span, a.jcs-JobTitle").first
                        title = (await title_el.inner_text(timeout=1500)).strip() if await title_el.count() else ""

                        company_el = card.locator("[data-testid='company-name'], span.companyName, .company").first
                        company = (await company_el.inner_text(timeout=1000)).strip() if await company_el.count() else ""

                        loc_el = card.locator("[data-testid='text-location'], .companyLocation, .location").first
                        location = (await loc_el.inner_text(timeout=1000)).strip() if await loc_el.count() else ""

                        link = ""
                        href = await title_el.get_attribute("href") if await title_el.count() else None
                        if href:
                            link = href if href.startswith("http") else f"https://www.indeed.com{href}"

                        if title:
                            jobs.append({
                                "Title": title,
                                "Company": company,
                                "Location": location,
                                "URL": link,
                            })
                    except Exception:
                        continue

            except PlaywrightTimeout:
                log.warning("Timeout on page %d", page_num + 1)
                break
            except Exception as e:
                log.error("Error on page %d: %s", page_num + 1, e)
                break

        await browser.close()

    if not jobs:
        log.warning("No jobs collected. Indeed may have blocked or changed layout.")
        return

    df = pd.DataFrame(jobs).drop_duplicates(subset=["URL"])
    df.to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")
    log.info("CSV saved → %s (%d rows)", OUTPUT_CSV, len(df))

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Indeed Jobs"
        headers = list(df.columns)
        fill = PatternFill("solid", fgColor="1F4E79")
        font = Font(bold=True, color="FFFFFF")
        thin = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.fill = fill
            cell.font = font
        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws.cell(r_idx, c_idx, val)
                cell.border = thin
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 55
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        wb.save(OUTPUT_XLSX)
        log.info("Excel saved → %s", OUTPUT_XLSX)
    except Exception as e:
        df.to_excel(OUTPUT_XLSX, index=False)
        log.warning("Basic Excel saved: %s", e)

    print("\n--- Preview ---")
    print(df.head().to_string(index=False))
    print(f"\nTotal unique jobs: {len(df)}")


if __name__ == "__main__":
    asyncio.run(scrape_indeed())
